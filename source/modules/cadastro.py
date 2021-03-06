import tkinter as tk
from PIL import Image, ImageTk
import pygame
import time as delay
from openpyxl import *
from modules import janelaconfig
import os
import tkinter.messagebox

# Função inicial destinada a verificar se a database já existe.

if os.path.exists('./database/database.xlsx'):
    database = load_workbook('./database/database.xlsx')
    formcadastro = database.active
else:
    database = Workbook()
    formcadastro = database.active


# Classe destinada ao formulário de cadastro.


class Formulario(tk.Toplevel):

    # Constroi o formulário

    def __init__(self):
        super().__init__()

        # Define propriedades do toplevel.
        Formulario.title(self, 'Ficha de Cadastro')
        Formulario.iconbitmap(self, './resources/img/icons/mainicon.ico')
        Formulario.configure(self, bg='#185c37')

        # Configura toplevel do formulário.
        conf_toplevel = janelaconfig.Janela(master=Formulario)
        conf_toplevel.centraliza_tamanho(janela=self)
        Formulario.grid_rowconfigure(self, 0, weight=1)
        Formulario.grid_columnconfigure(self, 0, weight=1)

        # Frame destinado ao fomulário
        frame_cadastro = tk.Frame(self, bg='#185c37')
        frame_cadastro.grid(sticky='', padx=120, pady=120)

        # Labels destinados a cada entry.
        label_nome = tk.Label(frame_cadastro,
                              text='Nome:',
                              font='Roboto 12 bold',
                              fg='#ffffff',
                              bg='#185c37',
                              justify='left',
                              anchor='w'
                              )

        label_nascimento = tk.Label(frame_cadastro,
                                    text='Data de Nascimento:',
                                    font='Roboto 12 bold',
                                    fg='#ffffff',
                                    bg='#185c37',
                                    justify='left',
                                    anchor='w'
                                    )

        label_atividade = tk.Label(frame_cadastro,
                                   text='Atividade:',
                                   font='Roboto 12 bold',
                                   fg='#ffffff',
                                   bg='#185c37',
                                   justify='left',
                                   anchor='w'
                                   )

        label_endereco = tk.Label(frame_cadastro,
                                  text='Endereço:',
                                  font='Roboto 12 bold',
                                  fg='#ffffff',
                                  bg='#185c37',
                                  justify='left',
                                  anchor='w'
                                  )

        label_telefone = tk.Label(frame_cadastro,
                                  text='Telefone:',
                                  font='Roboto 12 bold',
                                  fg='#ffffff',
                                  bg='#185c37',
                                  justify='left',
                                  anchor='w'
                                  )

        label_email = tk.Label(frame_cadastro,
                               text='E-mail:',
                               font='Roboto 12 bold',
                               fg='#ffffff',
                               bg='#185c37',
                               justify='left',
                               anchor='w'
                               )

        label_rg = tk.Label(frame_cadastro,
                            text='RG:',
                            font='Roboto 12 bold',
                            fg='#ffffff',
                            bg='#185c37',
                            justify='left',
                            anchor='w'
                            )

        label_cpf = tk.Label(frame_cadastro,
                             text='CPF:',
                             font='Roboto 12 bold',
                             fg='#ffffff',
                             bg='#185c37',
                             justify='left',
                             anchor='w'
                             )

        # Botões

        # Icone do Botão salvarform.
        icon_salvarformabre = Image.open('./resources/img/icons/salvaricon.ico')  # Carrega icone do botão.
        icon_salvarformabre = icon_salvarformabre.resize((30, 30), Image.ANTIALIAS)  # Redimensiona o icone.
        salvarform_icon = ImageTk.PhotoImage(icon_salvarformabre)  # Define icone do botão.

        icon_retornamenumabre = Image.open('./resources/img/icons/returnmenu.ico')  # Carrega icone do botão.
        icon_retornamenumabre = icon_retornamenumabre.resize((30, 30), Image.ANTIALIAS)  # Redimensiona o icone.
        retornamenu_icon = ImageTk.PhotoImage(icon_retornamenumabre)  # Define icone do botão.

        botao_salvarform = tk.Button(self,
                                     text='Salvar',
                                     compound='left',
                                     image=salvarform_icon,
                                     padx=10,
                                     bg='#d5fdec',
                                     font='Roboto 12 bold',
                                     width=120,
                                     height=50,
                                     bd=5,
                                     relief='raised',
                                     command=lambda: [save_form(), cadastra()]
                                     )
        botao_salvarform.image = salvarform_icon  # Chama icone junto ao botão.

        botao_retornamenu = tk.Button(self,
                                      text='Retornar',
                                      compound='left',
                                      image=retornamenu_icon,
                                      padx=10,
                                      bg='#d5fdec',
                                      font='Roboto 12 bold',
                                      width=120,
                                      height=50,
                                      bd=5,
                                      relief='raised',
                                      command=lambda: [retorna_menu(), self.destroy()]
                                      )
        botao_retornamenu.image = retornamenu_icon  # Chama icone junto ao botão.

        # Entrys referente a cada imput do formulário
        nome_campo = tk.Entry(frame_cadastro, width=80)
        nascimento_campo = tk.Entry(frame_cadastro, width=10)
        atividade_campo = tk.Entry(frame_cadastro, width=30)
        endereco_campo = tk.Entry(frame_cadastro, width=80)
        telefone_campo = tk.Entry(frame_cadastro, width=12)
        email_campo = tk.Entry(frame_cadastro, width=40)
        rg_campo = tk.Entry(frame_cadastro, width=11)
        cpf_campo = tk.Entry(frame_cadastro, width=11)

        # Layout

        # Labels
        label_nome.grid(row=0, column=0, sticky='e')
        label_nascimento.grid(row=1, column=0, sticky='e')
        label_atividade.grid(row=2, column=0, sticky='e')
        label_endereco.grid(row=3, column=0, sticky='e')
        label_telefone.grid(row=4, column=0, sticky='e')
        label_email.grid(row=5, column=0, sticky='e')
        label_rg.grid(row=6, column=0, sticky='e')
        label_cpf.grid(row=7, column=0, sticky='e')

        # Entrys
        nome_campo.grid(row=0, column=1, sticky='w')
        nascimento_campo.grid(row=1, column=1, sticky='w')
        atividade_campo.grid(row=2, column=1, sticky='w')
        endereco_campo.grid(row=3, column=1, sticky='w')
        telefone_campo.grid(row=4, column=1, sticky='w')
        email_campo.grid(row=5, column=1, sticky='w')
        rg_campo.grid(row=6, column=1, sticky='w')
        cpf_campo.grid(row=7, column=1, sticky='w')

        # Botões
        botao_salvarform.grid(row=8, column=1, sticky='se')
        botao_retornamenu.grid(row=8, column=0, sticky='sw')

        # Configuração de layout do formcadastro.

        formcadastro.column_dimensions['A'].width = 80
        formcadastro.column_dimensions['B'].width = 10
        formcadastro.column_dimensions['C'].width = 30
        formcadastro.column_dimensions['D'].width = 80
        formcadastro.column_dimensions['E'].width = 12
        formcadastro.column_dimensions['F'].width = 40
        formcadastro.column_dimensions['G'].width = 11
        formcadastro.column_dimensions['H'].width = 11

        # Escreve no formcadastro o posicionamento dos dados.

        formcadastro.cell(row=1, column=1).value = "Nome"
        formcadastro.cell(row=1, column=2).value = "Data de Nascimento"
        formcadastro.cell(row=1, column=3).value = "Atividade"
        formcadastro.cell(row=1, column=4).value = "Endereço"
        formcadastro.cell(row=1, column=5).value = "Telefone"
        formcadastro.cell(row=1, column=6).value = "Email"
        formcadastro.cell(row=1, column=7).value = "RG"
        formcadastro.cell(row=1, column=8).value = "CPF"

        # Função destinada a limpar as entry do formulário.

        def limpa_entry():
            nome_campo.delete(0, 'end')
            nascimento_campo.delete(0, 'end')
            atividade_campo.delete(0, 'end')
            endereco_campo.delete(0, 'end')
            telefone_campo.delete(0, 'end')
            email_campo.delete(0, 'end')
            rg_campo.delete(0, 'end')
            cpf_campo.delete(0, 'end')

        # Função destinada a mascarar e validar a data de nascimento.

        nascimento_campo.insert('end', '  /  /    ')

        def mascara_nascimento(event):  # Event chamado quando número é digitado.
            if any(letra.isalpha() for letra in nascimento_campo.get()):
                return tkinter.messagebox.showinfo('Atenção!', 'A data de nascimento deve conter somente numeros.',
                                                   parent=self)
            elif len(nascimento_campo.get()) > 12:
                return tkinter.messagebox.showinfo('Atenção!', 'A data de nascimento deve conter no máximo 8 numeros.',
                                                   parent=self)
            elif len(nascimento_campo.get()) == 2:
                nascimento_campo.insert('end', '/')
            elif len(nascimento_campo.get()) == 5:
                nascimento_campo.insert('end', '/')
            elif len(nascimento_campo.get()) == 11:
                nascimento_campo.delete(10, 'end')

        nascimento_campo.bind('<KeyRelease>', mascara_nascimento)

        # Função destinada a coletar os dados da interface e colocar no formulario.

        def cadastra():
            # Informa que existe algum campo em branco.
            if (nome_campo.get() == "" or
                    nascimento_campo.get() == "" or
                    atividade_campo.get() == "" or
                    endereco_campo.get() == "" or
                    telefone_campo.get() == "" or
                    email_campo.get() == "" or
                    rg_campo.get() == "" or
                    cpf_campo.get() == ""):

                # Exibe mensagem de campos faltantes.
                tkinter.messagebox.showinfo('Atenção', 'Por favor, preencha todos os campos.', parent=self)

            # Validação destinada aos campos.

            # Valida Nome:
            elif nome_campo.get().isnumeric() or len(nome_campo.get()) > 80:

                # Exibe mensagem de campo nome incorreto.
                tkinter.messagebox.showinfo('Atenção!', 'Por favor, o campo de nome dever conter apenas letras e '
                                                        'possuir no máximo 80 caracteres.', parent=self)

            # nacimento_campo validado na linha 233.

            # Valida Atividade:
            elif atividade_campo.get().isnumeric() or len(atividade_campo.get()) > 30:

                # Exibe mensagem de campo atividade incorreto.
                tkinter.messagebox.showinfo('Atenção!', 'Por favor, o campo de atividade dever conter apenas letras e '
                                                        'possuir no máximo 30 caracteres.', parent=self)

            # Valida Endereço:
            elif len(endereco_campo.get()) > 80:

                # Exibe mensagem de campo endereço incorreto.
                tkinter.messagebox.showinfo('Atenção!', 'Por favor, o campo de endereço dever possuir '
                                                        'no máximo 80 caracteres.', parent=self)

            # Valida Telefone:
            elif any(letra.isalpha() for letra in telefone_campo.get()):
                return tkinter.messagebox.showinfo('Atenção!', 'O telefone deve conter apenas números.',
                                                   parent=self)
            elif len(telefone_campo.get()) > 12:

                # Exibe mensagem de campo telefone incorreto.
                tkinter.messagebox.showinfo('Atenção!', 'O telefone deve possuir no máximo 12 digitos.',
                                            parent=self)

            # Valida E-mail:
            elif len(email_campo.get()) > 40:

                # Exibe mensagem de campo email incorreto.
                tkinter.messagebox.showinfo('Atenção!', 'O email deve possuir no máximo 40 caracteres.',
                                            parent=self)

            # Valida RG:
            elif any(letra.isalpha() for letra in rg_campo.get()):
                return tkinter.messagebox.showinfo('Atenção!', 'O RG deve conter apenas números.',
                                                   parent=self)
            elif len(rg_campo.get()) > 11:

                # Exibe mensagem de campo RG incorreto.
                tkinter.messagebox.showinfo('Atenção!', 'O RG deve possuir no máximo 11 digitos.',
                                            parent=self)

            # Valida CPF:
            elif any(letra.isalpha() for letra in cpf_campo.get()):
                return tkinter.messagebox.showinfo('Atenção!', 'O CPF deve conter apenas números.',
                                                   parent=self)
            elif len(cpf_campo.get()) > 11:

                # Exibe mensagem de campo RG incorreto.
                tkinter.messagebox.showinfo('Atenção!', 'O CPF deve possuir no máximo 11 digitos.',
                                            parent=self)

            else:
                # Define o maximo de linhas e colunas de acordo com o tamanho da database.
                linha_atual = formcadastro.max_row
                # coluna_atual = formcadastro.max_column

                # Coleta informações no formulário e preenche a database.
                formcadastro.cell(row=linha_atual + 1, column=1).value = nome_campo.get()
                formcadastro.cell(row=linha_atual + 1, column=2).value = nascimento_campo.get()
                formcadastro.cell(row=linha_atual + 1, column=3).value = atividade_campo.get()
                formcadastro.cell(row=linha_atual + 1, column=4).value = endereco_campo.get()
                formcadastro.cell(row=linha_atual + 1, column=5).value = telefone_campo.get()
                formcadastro.cell(row=linha_atual + 1, column=6).value = email_campo.get()
                formcadastro.cell(row=linha_atual + 1, column=7).value = rg_campo.get()
                formcadastro.cell(row=linha_atual + 1, column=8).value = cpf_campo.get()

                # Salva a database
                database.save('./database/database.xlsx')

                # Define o foco no campo nome.
                nome_campo.focus_set()

                # Limpa as entry boxes.
                limpa_entry()

                # Exibe mensagem de confirmação
                tkinter.messagebox.showinfo('Sucesso', 'Cadastro realizado com sucesso.', parent=self)

        # Define o foco inicial de preenchimento do formulário.
        nome_campo.focus_set()


# Funções Gerais

# Som do click no botão.


def click_sound():
    pygame.init()
    pygame.mixer.music.load('./resources/sound/mouseclick.wav')
    pygame.mixer.music.play()
    pygame.event.wait(timeout=1)


# Som de salvar formulário.


def save_sound():
    pygame.init()
    pygame.mixer.music.load('./resources/sound/saved.wav')
    pygame.mixer.music.play()
    pygame.event.wait(timeout=1)


# Realiza ações ao clicar.


def on_click():
    click_sound()


# Salva o formulário na database.


def save_form():
    save_sound()
    delay.sleep(0.1)


# Retorna a janela inicial.


def retorna_menu():
    on_click()
    delay.sleep(0.1)
